"""
exporter.py — Generate downloadable CSV and XLSX from scrape results.
"""

import io
import csv
from datetime import datetime

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from scraper import ScrapedResult, ErrorResult


RESULT_COLUMNS = ["#", "URL", "Vendor", "Product", "Edition", "Version", "Licence Metric", "EOS", "EOL"]
ERROR_COLUMNS  = ["#", "URL", "Error Reason"]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_border(color="D1D5DB"):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header(cell, bg_color, font_color="FFFFFF"):
    cell.font = Font(bold=True, color=font_color, size=11, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _make_border("9CA3AF")


def _auto_fit_columns(ws, min_width=12, max_width=55):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val_len = len(str(cell.value or ""))
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        adjusted = max(min_width, min(max_width, max_len + 4))
        ws.column_dimensions[col_letter].width = adjusted


# ---------------------------------------------------------------------------
# CSV Export
# ---------------------------------------------------------------------------

def to_csv(results: list[ScrapedResult], errors: list[ErrorResult]) -> bytes:
    """Build a CSV with results and errors in two sections."""
    buf = io.StringIO()
    writer = csv.writer(buf)

    writer.writerow(RESULT_COLUMNS)
    for i, r in enumerate(results, 1):
        writer.writerow([i, r.url, r.vendor, r.product, r.edition, r.version, r.licence_metric, r.eos, r.eol])

    writer.writerow([])
    writer.writerow(ERROR_COLUMNS)
    for i, e in enumerate(errors, 1):
        writer.writerow([i, e.url, e.error])

    return buf.getvalue().encode("utf-8-sig")


# ---------------------------------------------------------------------------
# XLSX Export — styled workbook (simple scrape results)
# ---------------------------------------------------------------------------

def to_xlsx(results: list[ScrapedResult], errors: list[ErrorResult]) -> bytes:
    """Build a styled XLSX with two sheets: Results and Errors."""
    wb = openpyxl.Workbook()

    ws_results = wb.active
    ws_results.title = "Scrape Results"
    ws_results.sheet_view.showGridLines = False
    ws_results.row_dimensions[1].height = 30

    for col_idx, col_name in enumerate(RESULT_COLUMNS, 1):
        cell = ws_results.cell(row=1, column=col_idx, value=col_name)
        _style_header(cell, "4F46E5")

    for i, r in enumerate(results, 1):
        row_num = i + 1
        ws_results.row_dimensions[row_num].height = 20
        data = [i, r.url, r.vendor, r.product, r.edition, r.version, r.licence_metric, r.eos, r.eol]
        alt = (i % 2 == 0)
        bg = "F9FAFB" if alt else "FFFFFF"
        for col_idx, value in enumerate(data, 1):
            cell = ws_results.cell(row=row_num, column=col_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _make_border("E5E7EB")
            cell.font = Font(size=10, name="Calibri")

    _auto_fit_columns(ws_results)
    ws_results.freeze_panes = "A2"

    ws_errors = wb.create_sheet(title="Error URLs")
    ws_errors.sheet_view.showGridLines = False
    ws_errors.row_dimensions[1].height = 30

    for col_idx, col_name in enumerate(ERROR_COLUMNS, 1):
        cell = ws_errors.cell(row=1, column=col_idx, value=col_name)
        _style_header(cell, "DC2626")

    for i, e in enumerate(errors, 1):
        row_num = i + 1
        ws_errors.row_dimensions[row_num].height = 20
        alt = (i % 2 == 0)
        bg = "F9FAFB" if alt else "FFFFFF"
        data = [i, e.url, e.error]
        for col_idx, value in enumerate(data, 1):
            cell = ws_errors.cell(row=row_num, column=col_idx, value=value)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _make_border("E5E7EB")
            cell.font = Font(size=10, name="Calibri")

    _auto_fit_columns(ws_errors)
    ws_errors.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Augmented XLSX Export
# The output has exactly:
#   - The 8 original input columns (as-is from the uploaded file)
#   - 7 scraped counterpart columns (Scraped Vendor, Scraped Product, etc.)
#   - Each scraped cell is highlighted GREEN if it matches the original, RED if not
# ---------------------------------------------------------------------------

# The 8 expected input columns (case-insensitive fuzzy match)
INPUT_COL_KEYS = ["url", "vendor", "product", "edition", "version", "licence metric", "eos", "eol"]

# Labels for scraped counterpart columns
SCRAPED_LABELS = [
    "Scraped Vendor",
    "Scraped Product",
    "Scraped Edition",
    "Scraped Version",
    "Scraped Licence Metric",
    "Scraped EOS",
    "Scraped EOL",
]

# Which original column (by INPUT_COL_KEYS index) each scraped column compares to
# Index 0 = url (no scraped counterpart for url itself)
# Scraped Vendor -> compares to index 1 (vendor)
# Scraped Product -> compares to index 2 (product)
# ...
SCRAPED_COMPARE_TO = [1, 2, 3, 4, 5, 6, 7]  # indices into orig_cols list


def _fuzzy_find_col(df_columns, key):
    """Find a column in df by case-insensitive substring match."""
    key_lower = key.lower()
    for c in df_columns:
        if c.lower() == key_lower:
            return c
    for c in df_columns:
        if key_lower in c.lower() or c.lower() in key_lower:
            return c
    return None


def _safe_str(val):
    import pandas as pd
    if val is None:
        return ""
    try:
        if pd.isna(val):
            return ""
    except Exception:
        pass
    return str(val).strip()


def _is_match(orig_val: str, scraped_val: str) -> bool:
    """Return True if there is a meaningful match between the two values."""
    if not orig_val or not scraped_val or scraped_val in ("N/A", ""):
        return False
    a = orig_val.lower()
    b = scraped_val.lower()
    return a in b or b in a


def to_xlsx_augmented(df, orig_col_map: dict = None) -> bytes:
    """
    Build a styled XLSX from an augmented DataFrame.

    orig_col_map: dict mapping logical key -> actual df column name
        keys: url, vendor, product, edition, version, licence_metric, eos, eol
    scraped cols expected in df:
        vendor_webscraped, product_webscraped, edition_webscraped,
        version_webscraped, licence_metric_webscraped, eos_webscraped, eol_webscraped
    """
    import pandas as pd

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.sheet_view.showGridLines = False

    # -----------------------------------------------------------------------
    # Determine original column mapping
    # -----------------------------------------------------------------------
    if orig_col_map is None:
        orig_col_map = {}
        key_hints = {
            "url":            ["url", "link", "uri"],
            "vendor":         ["vendor"],
            "product":        ["product"],
            "edition":        ["edition"],
            "version":        ["version"],
            "licence_metric": ["licence metric", "license metric", "licensemetric", "metric"],
            "eos":            ["eos", "end of support"],
            "eol":            ["eol", "end of life"],
        }
        for key, hints in key_hints.items():
            for hint in hints:
                col = _fuzzy_find_col(df.columns, hint)
                if col:
                    orig_col_map[key] = col
                    break

    # Build the ordered list of (header_label, df_col_name) for original cols
    orig_display = [
        ("URL",            orig_col_map.get("url")),
        ("Vendor",         orig_col_map.get("vendor")),
        ("Product",        orig_col_map.get("product")),
        ("Edition",        orig_col_map.get("edition")),
        ("Version",        orig_col_map.get("version")),
        ("Licence Metric", orig_col_map.get("licence_metric")),
        ("EOS",            orig_col_map.get("eos")),
        ("EOL",            orig_col_map.get("eol")),
    ]

    # Scraped columns in df
    scraped_df_cols = [
        "vendor_webscraped",
        "product_webscraped",
        "edition_webscraped",
        "version_webscraped",
        "licence_metric_webscraped",
        "eos_webscraped",
        "eol_webscraped",
    ]

    # Which orig col each scraped col compares to (by orig_display index)
    # vendor_webscraped compares to orig_display[1] (Vendor), etc.
    scraped_compare_idx = [1, 2, 3, 4, 5, 6, 7]

    # -----------------------------------------------------------------------
    # Header row
    # -----------------------------------------------------------------------
    ws.row_dimensions[1].height = 34

    col_idx = 1
    # Original columns — dark blue header
    for label, _ in orig_display:
        cell = ws.cell(row=1, column=col_idx, value=label)
        _style_header(cell, "3730A3")  # indigo-800
        col_idx += 1

    # Divider visual (we add a narrow spacer column)
    spacer_col = col_idx
    ws.column_dimensions[get_column_letter(spacer_col)].width = 2
    cell = ws.cell(row=1, column=col_idx, value="")
    cell.fill = PatternFill("solid", fgColor="1E1B4B")
    col_idx += 1

    # Scraped columns — teal header
    for label in SCRAPED_LABELS:
        cell = ws.cell(row=1, column=col_idx, value=label)
        _style_header(cell, "0F766E")  # teal-700
        col_idx += 1

    total_cols = col_idx - 1

    # -----------------------------------------------------------------------
    # Data rows
    # -----------------------------------------------------------------------
    GREEN_LIGHT = "BBFBD0"
    RED_LIGHT   = "FECACA"
    GREEN_DARK  = "16A34A"
    RED_DARK    = "DC2626"
    ALT_BG      = "F8FAFC"
    NORM_BG     = "FFFFFF"

    for row_idx, (_, row) in enumerate(df.iterrows(), start=1):
        excel_row = row_idx + 1
        ws.row_dimensions[excel_row].height = 22
        alt = (row_idx % 2 == 0)
        default_bg = ALT_BG if alt else NORM_BG

        col_idx = 1

        # --- Write original columns ---
        orig_values = []
        for label, df_col in orig_display:
            val = _safe_str(row.get(df_col, "") if df_col else "")
            orig_values.append(val)
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.fill = PatternFill("solid", fgColor=default_bg)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = _make_border("E2E8F0")
            cell.font = Font(size=10, name="Calibri", color="1E293B")
            col_idx += 1

        # --- Spacer ---
        cell = ws.cell(row=excel_row, column=col_idx, value="")
        cell.fill = PatternFill("solid", fgColor="E0E7FF")
        col_idx += 1

        # --- Write scraped columns with green/red highlight ---
        for s_idx, s_col in enumerate(scraped_df_cols):
            scraped_val = _safe_str(row.get(s_col, ""))
            orig_idx = scraped_compare_idx[s_idx]
            orig_val = orig_values[orig_idx] if orig_idx < len(orig_values) else ""

            matched = _is_match(orig_val, scraped_val)
            has_value = bool(scraped_val and scraped_val not in ("N/A", ""))

            if not has_value:
                bg = default_bg
                fc = "94A3B8"  # slate-400
            elif matched:
                bg = GREEN_LIGHT
                fc = GREEN_DARK
            else:
                bg = RED_LIGHT
                fc = RED_DARK

            cell = ws.cell(row=excel_row, column=col_idx, value=scraped_val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = _make_border("E2E8F0")
            cell.font = Font(size=10, name="Calibri", color=fc, bold=matched)
            col_idx += 1

    # -----------------------------------------------------------------------
    # Auto-fit & freeze
    # -----------------------------------------------------------------------
    _auto_fit_columns(ws)
    # Keep spacer col narrow
    ws.column_dimensions[get_column_letter(spacer_col)].width = 2
    ws.freeze_panes = "A2"

    # -----------------------------------------------------------------------
    # Sheet 2 — Validation Score Report
    # -----------------------------------------------------------------------
    ws2 = wb.create_sheet(title="Validation Report")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "7C3AED"

    FIELD_LABELS = ["Vendor", "Product", "Edition", "Version", "Licence Metric", "EOS", "EOL"]

    # --- Sheet 2 Header banner ---
    ws2.merge_cells("A1:J1")
    banner = ws2.cell(row=1, column=1, value="✦  Validation Score Report")
    banner.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    banner.fill = PatternFill("solid", fgColor="3730A3")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:J2")
    sub = ws2.cell(row=2, column=1, value="Compares scraped values against original input data. Score = matched fields / total comparable fields.")
    sub.font = Font(size=9, color="94A3B8", italic=True, name="Calibri")
    sub.fill = PatternFill("solid", fgColor="1E1B4B")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 20

    # --- Per-row table header (row 4) ---
    row_tbl_headers = ["#", "URL / Row ID", "Vendor", "Product", "Edition", "Version",
                       "Licence Metric", "EOS", "EOL", "Score", "Rating"]
    ws2.row_dimensions[4].height = 28
    hdr_bg  = "0F766E"
    for ci, h in enumerate(row_tbl_headers, 1):
        cell = ws2.cell(row=4, column=ci, value=h)
        _style_header(cell, hdr_bg)

    # Per-row data
    row_scores = []   # list of (match_count, total_comparable) per df row

    for row_idx, (_, row) in enumerate(df.iterrows(), start=1):
        excel_row = row_idx + 4
        ws2.row_dimensions[excel_row].height = 20
        alt = (row_idx % 2 == 0)
        default_bg2 = "F8FAFC" if alt else "FFFFFF"

        # Row number
        cell = ws2.cell(row=excel_row, column=1, value=row_idx)
        cell.fill = PatternFill("solid", fgColor=default_bg2)
        cell.font = Font(size=9, color="64748B", name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _make_border("E2E8F0")

        # URL / Row ID (use url col if available)
        url_col_name = orig_col_map.get("url") if orig_col_map else None
        url_val = _safe_str(row.get(url_col_name, f"Row {row_idx}") if url_col_name else f"Row {row_idx}")
        cell = ws2.cell(row=excel_row, column=2, value=url_val)
        cell.fill = PatternFill("solid", fgColor=default_bg2)
        cell.font = Font(size=9, color="1E293B", name="Calibri")
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = _make_border("E2E8F0")

        # Per-field match cells
        orig_display_local = [
            orig_col_map.get("vendor") if orig_col_map else None,
            orig_col_map.get("product") if orig_col_map else None,
            orig_col_map.get("edition") if orig_col_map else None,
            orig_col_map.get("version") if orig_col_map else None,
            orig_col_map.get("licence_metric") if orig_col_map else None,
            orig_col_map.get("eos") if orig_col_map else None,
            orig_col_map.get("eol") if orig_col_map else None,
        ]

        field_match_flags = []
        for fi, (s_col, o_col) in enumerate(zip(scraped_df_cols, orig_display_local)):
            scraped_v = _safe_str(row.get(s_col, ""))
            orig_v    = _safe_str(row.get(o_col, "") if o_col else "")

            has_scraped = bool(scraped_v and scraped_v not in ("N/A", "") and not scraped_v.startswith("ERROR"))
            matched     = _is_match(orig_v, scraped_v) if has_scraped else False

            if has_scraped:
                field_match_flags.append(1 if matched else 0)
                symbol = "✓" if matched else "✗"
                bg = "BBFBD0" if matched else "FECACA"
                fc = "16A34A" if matched else "DC2626"
            else:
                field_match_flags.append(None)  # not comparable
                symbol = "—"
                bg = default_bg2
                fc = "94A3B8"

            cell = ws2.cell(row=excel_row, column=3 + fi, value=symbol)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=10, bold=has_scraped, color=fc, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _make_border("E2E8F0")

        # Score calculation
        comparable = [f for f in field_match_flags if f is not None]
        matched_cnt = sum(comparable)
        total_cnt   = len(comparable)
        score_pct   = round((matched_cnt / total_cnt) * 100) if total_cnt > 0 else 0
        row_scores.append((matched_cnt, total_cnt, score_pct))

        # Score cell
        score_label = f"{matched_cnt}/{total_cnt}  ({score_pct}%)"
        cell = ws2.cell(row=excel_row, column=10, value=score_label)
        if score_pct >= 80:
            s_bg, s_fc = "BBFBD0", "16A34A"
        elif score_pct >= 50:
            s_bg, s_fc = "FEF9C3", "92400E"
        else:
            s_bg, s_fc = "FECACA", "DC2626"
        cell.fill = PatternFill("solid", fgColor=s_bg)
        cell.font = Font(size=10, bold=True, color=s_fc, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _make_border("E2E8F0")

        # Rating label
        if score_pct >= 80:
            rating, r_bg, r_fc = "PASS ✓", "BBFBD0", "16A34A"
        elif score_pct >= 50:
            rating, r_bg, r_fc = "REVIEW ⚠", "FEF9C3", "92400E"
        else:
            rating, r_bg, r_fc = "FAIL ✗", "FECACA", "DC2626"

        cell = ws2.cell(row=excel_row, column=11, value=rating)
        cell.fill = PatternFill("solid", fgColor=r_bg)
        cell.font = Font(size=10, bold=True, color=r_fc, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _make_border("E2E8F0")

    # -----------------------------------------------------------------------
    # Summary panel (below the per-row table, with a gap)
    # -----------------------------------------------------------------------
    total_rows    = len(row_scores)
    summary_start = total_rows + 6  # row offset for summary block

    # Summary title
    ws2.merge_cells(f"A{summary_start}:K{summary_start}")
    cell = ws2.cell(row=summary_start, column=1, value="📊  Overall Summary")
    cell.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _make_border("334155")
    ws2.row_dimensions[summary_start].height = 28

    # Overall stats
    all_matched = sum(s[0] for s in row_scores)
    all_total   = sum(s[1] for s in row_scores)
    overall_pct = round((all_matched / all_total) * 100) if all_total > 0 else 0
    pass_rows   = sum(1 for s in row_scores if s[2] >= 80)
    review_rows = sum(1 for s in row_scores if 50 <= s[2] < 80)
    fail_rows   = sum(1 for s in row_scores if s[2] < 50)

    summary_kv = [
        ("Total Rows Processed",   total_rows),
        ("Total Fields Compared",  all_total),
        ("Total Fields Matched",   all_matched),
        ("Overall Match Rate",     f"{overall_pct}%"),
        ("Rows: PASS (≥80%)",     pass_rows),
        ("Rows: REVIEW (50–79%)", review_rows),
        ("Rows: FAIL (<50%)",     fail_rows),
    ]

    for i, (label, val) in enumerate(summary_kv):
        r = summary_start + 1 + i
        ws2.row_dimensions[r].height = 20
        lc = ws2.cell(row=r, column=1, value=label)
        lc.font = Font(size=10, bold=True, color="E2E8F0", name="Calibri")
        lc.fill = PatternFill("solid", fgColor="1E293B")
        lc.alignment = Alignment(vertical="center")
        lc.border = _make_border("334155")
        ws2.merge_cells(f"A{r}:C{r}")

        vc = ws2.cell(row=r, column=4, value=val)
        vc.font = Font(size=10, bold=True, color="7DD3FC", name="Calibri")
        vc.fill = PatternFill("solid", fgColor="0F172A")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = _make_border("334155")

    # Per-field match rate table
    field_tbl_start = summary_start + len(summary_kv) + 2
    ws2.row_dimensions[field_tbl_start].height = 26

    cell = ws2.cell(row=field_tbl_start, column=1, value="Field")
    _style_header(cell, "4F46E5")
    ws2.merge_cells(f"A{field_tbl_start}:C{field_tbl_start}")

    cell = ws2.cell(row=field_tbl_start, column=4, value="Matched")
    _style_header(cell, "4F46E5")

    cell = ws2.cell(row=field_tbl_start, column=5, value="Compared")
    _style_header(cell, "4F46E5")

    cell = ws2.cell(row=field_tbl_start, column=6, value="Match Rate")
    _style_header(cell, "4F46E5")

    # Compute per-field stats
    for fi, field_label in enumerate(FIELD_LABELS):
        s_col = scraped_df_cols[fi]
        o_col_key = ["vendor","product","edition","version","licence_metric","eos","eol"][fi]
        o_col = orig_col_map.get(o_col_key) if orig_col_map else None

        f_matched = 0
        f_total   = 0
        for _, row in df.iterrows():
            sv = _safe_str(row.get(s_col, ""))
            ov = _safe_str(row.get(o_col, "") if o_col else "")
            if sv and sv not in ("N/A", "") and not sv.startswith("ERROR"):
                f_total += 1
                if _is_match(ov, sv):
                    f_matched += 1

        f_pct = round((f_matched / f_total) * 100) if f_total > 0 else 0
        frow  = field_tbl_start + 1 + fi
        ws2.row_dimensions[frow].height = 20

        lc = ws2.cell(row=frow, column=1, value=field_label)
        lc.font = Font(size=10, bold=True, color="1E293B", name="Calibri")
        lc.fill = PatternFill("solid", fgColor="EEF2FF")
        lc.alignment = Alignment(vertical="center")
        lc.border = _make_border("E2E8F0")
        ws2.merge_cells(f"A{frow}:C{frow}")

        mc = ws2.cell(row=frow, column=4, value=f_matched)
        mc.font = Font(size=10, color="16A34A", name="Calibri")
        mc.fill = PatternFill("solid", fgColor="F0FFF4")
        mc.alignment = Alignment(horizontal="center", vertical="center")
        mc.border = _make_border("E2E8F0")

        tc = ws2.cell(row=frow, column=5, value=f_total)
        tc.font = Font(size=10, color="334155", name="Calibri")
        tc.fill = PatternFill("solid", fgColor="F8FAFC")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        tc.border = _make_border("E2E8F0")

        bar = f"{f_pct}%"
        pc = ws2.cell(row=frow, column=6, value=bar)
        if f_pct >= 80:
            p_bg, p_fc = "BBFBD0", "16A34A"
        elif f_pct >= 50:
            p_bg, p_fc = "FEF9C3", "92400E"
        else:
            p_bg, p_fc = "FECACA", "DC2626"
        pc.fill = PatternFill("solid", fgColor=p_bg)
        pc.font = Font(size=10, bold=True, color=p_fc, name="Calibri")
        pc.alignment = Alignment(horizontal="center", vertical="center")
        pc.border = _make_border("E2E8F0")

    # Auto-fit Sheet 2
    _auto_fit_columns(ws2, min_width=10, max_width=50)
    ws2.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
